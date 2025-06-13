import os
import io
import base64
import json
import uvicorn
import socketio
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
import fitz  # PyMuPDF for PDF text extraction
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import httpx # Using httpx for LM Studio API calls
import re # For regex-based JSON extraction
import traceback # For detailed traceback logging
from dotenv import load_dotenv # To load .env file

# NEW IMPORTS for document parsing
from docx import Document # For .docx files
import openpyxl # For .xlsx files

# Load environment variables from .env file
load_dotenv()

# Initialize FastAPI app
app = FastAPI()

# Configure CORS middleware
# Allow all origins for development. In production, restrict this to your frontend's domain.
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], # In production, change this to your frontend's domain, e.g., ["https://your-laravel-app.com"]
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Initialize Socket.IO server
sio = socketio.AsyncServer(async_mode='asgi', cors_allowed_origins='*') # In production, change this to your frontend's domain
# Wrap FastAPI application with Socket.IO ASGI app
app_with_sio = socketio.ASGIApp(sio, app)

# Placeholder for API key - will be provided by the Canvas environment at runtime
# Not strictly needed if only using local Ollama, but kept for consistency if switching to cloud later
API_KEY = os.environ.get("GEMINI_API_KEY", "") # Fallback to empty string if not set

# Define a constant for the literal JSON error message to avoid f-string parsing issues
JSON_ERROR_RESPONSE_LITERAL = '{"error": "JSON output malformed."}'

# Define the model name globally for access throughout the application
# LM Studio uses the model name as configured within LM Studio, typically the file name.
GLOBAL_MODEL_NAME = "gemma3:4b" # Still using gemma3:4b as per your last instruction

# LM Studio API endpoint (default for OpenAI-compatible server)
# Ensure this matches the IP/Port shown in LM Studio's Local Inference Server tab.
LM_STUDIO_API_URL = "http://192.168.234.1:1234/v1/chat/completions"

# File size limit (10 MB)
MAX_FILE_SIZE_BYTES = 10 * 1024 * 1024 # 10 MB

# Allowed MIME types
ALLOWED_MIME_TYPES = {
    'application/pdf': '.pdf',
    'application/vnd.openxmlformats-officedocument.wordprocessingml.document': '.docx',
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': '.xlsx',
}

# --- Utility Functions for Text Extraction ---

def extract_text_from_pdf(pdf_bytes):
    """
    Extracts text content from PDF bytes using PyMuPDF.
    """
    text_content = ""
    try:
        # Open the PDF document from bytes
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        for page_num in range(len(doc)):
            page = doc.load_page(page_num)
            text_content += page.get_text()
        doc.close()
    except Exception as e:
        print(f"Error extracting text from PDF: {e}")
        return None
    return text_content

def extract_text_from_docx(docx_bytes):
    """
    Extracts text content from DOCX bytes using python-docx.
    """
    text_content = ""
    try:
        doc = Document(io.BytesIO(docx_bytes))
        for paragraph in doc.paragraphs:
            text_content += paragraph.text + "\n"
    except Exception as e:
        print(f"Error extracting text from DOCX: {e}")
        return None
    return text_content

def extract_text_from_xlsx(xlsx_bytes):
    """
    Extracts text content from XLSX bytes using openpyxl.
    Iterates through all sheets and cells.
    """
    text_content = ""
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text_content += f"--- Sheet: {sheet_name} ---\n"
            for row in sheet.iter_rows():
                row_values = []
                for cell in row:
                    if cell.value is not None:
                        row_values.append(str(cell.value))
                if row_values:
                    text_content += " ".join(row_values) + "\n"
            text_content += "\n" # Add a newline after each sheet
    except Exception as e:
        print(f"Error extracting text from XLSX: {e}")
        return None
    return text_content

def generate_pdf_from_text(text_content):
    """
    Generates a new PDF from a given string of text using ReportLab.
    The content is wrapped in a Paragraph for better text flow.
    """
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    style_normal = styles['Normal']

    # Define content area and line height
    left_margin = inch
    top_margin = letter[1] - inch
    content_width = letter[0] - 2 * inch
    line_height = 14  # Approximate line height for text

    # Split text into paragraphs for better rendering
    paragraphs = text_content.split('\n')
    current_y = top_margin

    for para_text in paragraphs:
        # Create a Paragraph object
        paragraph = Paragraph(para_text, style_normal)

        # Calculate height required for the paragraph
        # This is a simplification; for complex layouts, flowables would be used
        text_lines = paragraph.wrapOn(p, content_width, 0)[1] / line_height
        paragraph_height = text_lines * line_height

        # Check if new page is needed
        if current_y - paragraph_height < inch: # If less than 1 inch from bottom
            p.showPage()
            current_y = top_margin

        # Draw the paragraph
        paragraph.drawOn(p, left_margin, current_y - paragraph_height)
        current_y -= paragraph_height + 6 # Add some spacing between paragraphs

    p.save()
    buffer.seek(0)
    return buffer.getvalue()

async def call_lm_studio_api(prompt: str, response_schema: dict = None):
    """
    Makes a call to LM Studio's OpenAI-compatible API for the specified model.
    Handles cases where the model might output multiple concatenated JSON objects.
    """
    model_name = GLOBAL_MODEL_NAME
    raw_response_content = None # To store raw response text from LM Studio
    
    try:
        messages = [{"role": "user", "content": prompt}]

        # LM Studio API request payload (OpenAI compatible)
        payload = {
            "model": model_name,
            "messages": messages,
            "temperature": 0.7,   # Controls randomness; lower for more deterministic
            "top_p": 0.7,         # Nucleus sampling; controls diversity - UPDATED
            "max_tokens": 4096,   # Max new tokens to predict - UPDATED
            # "stop": ["```"]       # Stop generation if "```" is encountered - COMMENTED OUT
        }

        async with httpx.AsyncClient(timeout=180.0) as client: # Added client timeout
            response = await client.post(LM_STUDIO_API_URL, json=payload)
            response.raise_for_status() # Raise an exception for HTTP errors (4xx or 5xx)

            full_lm_studio_response = response.json()
            # Safely get the content. LM Studio (OpenAI-compatible) structure.
            raw_response_content = full_lm_studio_response.get('choices', [{}])[0].get('message', {}).get('content', '')

        if not raw_response_content:
            print(f"ERROR: LM Studio model {model_name} returned empty content. Full LM Studio response: {full_lm_studio_response}")
            return {"error": f"LM Studio model {model_name} returned an empty response. This might indicate a problem with the model or insufficient resources.", "raw_response": str(full_lm_studio_response)}
        
        # Check if the raw response content is the literal error string
        if raw_response_content.strip() == JSON_ERROR_RESPONSE_LITERAL:
            print(f"ERROR: LM Studio model {model_name} explicitly returned JSON_ERROR_RESPONSE_LITERAL.")
            return {"error": "AI Model explicitly indicated malformed JSON output.", "raw_response": raw_response_content}


        # Step 1: Remove markdown code block wrappers if present
        result_text = raw_response_content
        if result_text.startswith('```json') and result_text.endswith('```'):
            result_text = result_text.removeprefix('```json').removesuffix('```').strip()
        elif result_text.startswith('```') and result_text.endswith('```'):
            result_text = result_text.removeprefix('```').removesuffix('```').strip()

        # If structured response is expected, try to parse it.
        if response_schema:
            final_result = {}
            decoder = json.JSONDecoder()
            idx = 0
            while idx < len(result_text):
                try:
                    obj, new_idx = decoder.raw_decode(result_text[idx:])
                    if isinstance(obj, dict):
                        final_result.update(obj)
                    idx += new_idx
                    # Skip any whitespace/newlines between JSON objects
                    while idx < len(result_text) and result_text[idx].isspace():
                        idx += 1
                except json.JSONDecodeError as e:
                    print(f"DEBUG: JSONDecodeError at index {idx}. Error: {e}. Raw text segment causing error: '{result_text[idx:idx+50]}...'")
                    next_brace = result_text.find('{', idx)
                    if next_brace != -1:
                        idx = next_brace
                    else:
                        break # No more '{' found, stop trying to parse
                except Exception as e:
                    print(f"DEBUG: Unexpected error during JSON decoding loop: {e}. Raw text: {result_text}")
                    traceback.print_exc()
                    break 

            if not final_result: # If nothing was parsed successfully
                print(f"WARNING: Expected JSON from {model_name}, but failed to parse and merge any valid JSON objects. Raw response after cleaning: '{result_text}'")
                return {"error": f"Failed to parse JSON from {model_name} API. Ensure {model_name} is outputting valid JSON or can be merged.", "raw_response": result_text}
            
            return final_result
        else:
            # If no specific schema is expected, return raw text
            return result_text

    except httpx.TimeoutException:
        error_message = f"LM Studio API call timed out after 180 seconds for model {model_name}. The model might be taking too long to generate a response or is stuck."
        print(f"ERROR: {error_message}")
        traceback.print_exc()
        return {"error": error_message, "raw_response": "Timeout"}
    except httpx.HTTPStatusError as e:
        error_message = f"HTTP error from LM Studio API for model {model_name}: {e.response.status_code} - {e.response.text}"
        print(f"ERROR: {error_message}")
        traceback.print_exc()
        return {"error": error_message, "raw_response": e.response.text}
    except httpx.RequestError as e:
        error_message = f"Network or request error connecting to LM Studio API for model {model_name}: {e}"
        print(f"ERROR: {error_message}")
        traceback.print_exc()
        return {"error": error_message, "raw_response": str(e)}
    except Exception as e:
        error_message = f"General Exception during LM Studio API call for {model_name}: {e}"
        print(f"ERROR: {error_message}")
        traceback.print_exc()
        return {"error": error_message, "raw_response": str(e)}
    finally:
        print(f"INFO: Finished call_lm_studio_api for {model_name}.")
        if raw_response_content:
            print(f"INFO: Full raw LM Studio content (first 200 chars): {str(raw_response_content)[:200]}...")
        else:
            print("INFO: No raw LM Studio content captured or content was empty.")


# --- Socket.IO Event Handlers ---

@sio.event
async def connect(sid, environ, auth):
    """Handles client connection."""
    print(f'Client connected: {sid}')
    await sio.emit('status', {'message': 'Connected to backend!'}, room=sid)

@sio.event
async def disconnect(sid):
    """Handles client disconnection."""
    print(f'Client disconnected: {sid}')

@sio.event
async def upload_resume_and_jd(sid, data):
    """
    Receives resume PDF and job description, then performs analysis.
    Handles multiple file types and size validation.
    """
    print(f"Received upload_resume_and_jd event from {sid}")
    resume_file_b64 = data.get('resume_file')
    job_description = data.get('job_description')
    file_type = data.get('file_type') # Get file type from frontend

    if not resume_file_b64 or not job_description or not file_type:
        await sio.emit('error', {'message': 'Resume file, Job Description, and file type are required.'}, room=sid)
        return

    try:
        # Decode base64 file to bytes
        resume_bytes = base64.b64decode(resume_file_b64)

        # Server-side file size validation
        if len(resume_bytes) > MAX_FILE_SIZE_BYTES:
            await sio.emit('error', {'message': f'File size exceeds the limit of {MAX_FILE_SIZE_BYTES / (1024 * 1024):.0f} MB.'}, room=sid)
            return

        resume_text = None
        if file_type == 'application/pdf':
            resume_text = extract_text_from_pdf(resume_bytes)
        elif file_type == 'application/vnd.openxmlformats-officedocument.wordprocessingml.document':
            resume_text = extract_text_from_docx(resume_bytes)
        elif file_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
            resume_text = extract_text_from_xlsx(resume_bytes)
        else:
            await sio.emit('error', {'message': 'Unsupported file type. Please upload PDF, DOCX, or XLSX.'}, room=sid)
            return

        if not resume_text:
            await sio.emit('error', {'message': f'Could not extract text from the provided {file_type} file. It might be empty or corrupted.'}, room=sid)
            return

        # Prompt for resume analysis
        analysis_prompt = f"""
As an expert resume analyzer, compare the following resume to the provided job description.
Your task is to:
1.  Assign a compatibility score between 1 and 10 (10 being a perfect match).
2.  Identify key areas where the resume could be improved to better align with the job description. Provide specific, actionable suggestions.
3.  Suggest modifications to the resume content, focusing on incorporating relevant keywords, emphasizing achievements pertinent to the role, and making the resume more impactful for this specific job.

**IMPORTANT:** Your entire response MUST be a SINGLE JSON object. Do NOT include any other text, explanations, or markdown formatting.
The JSON object MUST have the following keys:
-   `score`: An integer from 1 to 10.
-   `suggestions`: An array of strings, each string being an actionable suggestion.
-   `revised_summary`: A brief (2-3 sentences) suggested professional summary/objective for the resume, tailored to the job description.
If you cannot generate valid JSON, provide the literal string '{JSON_ERROR_RESPONSE_LITERAL}'.

Resume:
```
{resume_text}
```

Job Description:
```
{job_description}
```
"""
        # Define JSON schema for structured response (for backend parsing, not directly for Ollama)
        analysis_schema = {
            "type": "OBJECT",
            "properties": {
                "score": {"type": "INTEGER", "description": "Compatibility score (1-10)"},
                "suggestions": {"type": "ARRAY", "items": {"type": "STRING"}},
                "revised_summary": {"type": "STRING", "description": "Suggested professional summary/objective"}
            },
            "required": ["score", "suggestions", "revised_summary"]
        }

        # Call LM Studio API for analysis
        print(f"Calling {GLOBAL_MODEL_NAME} via LM Studio for resume analysis...")
        analysis_result = await call_lm_studio_api(analysis_prompt, analysis_schema)

        if "error" in analysis_result:
            await sio.emit('error', {'message': f"AI Analysis Error: {analysis_result['error']}"}, room=sid)
            return

        print("Emitting analysis_result to client.")
        await sio.emit('analysis_result', {
            'score': analysis_result.get('score', 0),
            'suggestions': analysis_result.get('suggestions', []),
            'revised_summary': analysis_result.get('revised_summary', ''),
            'extracted_resume_text': resume_text # Send extracted text back for editing
        }, room=sid)

    except Exception as e:
        print(f"ERROR: Exception during upload_resume_and_jd: {e}")
        # Print full traceback for any exception occurring in this handler
        traceback.print_exc()
        await sio.emit('error', {'message': f'Server error during analysis: {e}'}, room=sid)

@sio.event
async def apply_edits(sid, data):
    """
    Receives edited resume text and generates a new PDF.
    """
    print(f"Received apply_edits event from {sid}")
    edited_resume_text = data.get('edited_resume_text')

    if not edited_resume_text:
        await sio.emit('error', {'message': 'Edited resume text is empty.'}, room=sid)
        return

    try:
        # Generate new PDF from edited text
        new_pdf_bytes = generate_pdf_from_text(edited_resume_text)
        new_pdf_b64 = base64.b64encode(new_pdf_bytes).decode('utf-8')

        print("Emitting updated_resume_pdf to client.")
        await sio.emit('updated_resume_pdf', {'pdf_b64': new_pdf_b64}, room=sid)

    except Exception as e:
        print(f"ERROR: Exception during apply_edits: {e}")
        traceback.print_exc()
        await sio.emit('error', {'message': f'Server error during PDF generation: {e}'}, room=sid)

@sio.event
async def request_interview_prep(sid, data):
    """
    Generates interview preparation materials based on job description.
    """
    print(f"Received request_interview_prep event from {sid}")
    job_description = data.get('job_description')

    if not job_description:
        await sio.emit('error', {'message': 'Job Description is required for interview prep.'}, room=sid)
        return

    try:
        # Prompt for interview preparation
        interview_prompt = f"""
As an expert interview coach, based on the following job description, generate:
1.  5-7 common interview questions that an applicant for this role might face.
2.  For each question, provide a brief, high-level outline of an ideal answer, highlighting what aspects the interviewer is looking for.
3.  3 general interview tips specifically relevant to preparing for an interview for this type of role.

**IMPORTANT:** Your entire response MUST be a SINGLE JSON object. Do NOT include any other text, explanations, or markdown formatting.
The JSON object MUST have the following keys:
-   `interview_questions`: An array of objects, each with `question` (string) and `ideal_answer_outline` (string).
-   `interview_tips`: An array of strings, each string being a general interview tip.
If you cannot generate valid JSON, provide the literal string '{JSON_ERROR_RESPONSE_LITERAL}'.

Job Description:
```
{job_description}
```
"""
        # Define JSON schema for structured response (for backend parsing, not directly for Ollama)
        interview_schema = {
            "type": "OBJECT",
            "properties": {
                "interview_questions": {
                    "type": "ARRAY",
                    "items": {
                        "type": "OBJECT",
                        "properties": {
                            "question": {"type": "STRING"},
                            "ideal_answer_outline": {"type": "STRING"}
                        },
                        "required": ["question", "ideal_answer_outline"]
                    }
                },
                "interview_tips": {"type": "ARRAY", "items": {"type": "STRING"}}
            },
            "required": ["interview_questions", "interview_tips"]
        }

        # Call LM Studio API for interview prep
        print(f"Calling {GLOBAL_MODEL_NAME} via LM Studio for interview prep...")
        prep_result = await call_lm_studio_api(interview_prompt, interview_schema)

        if "error" in prep_result:
            await sio.emit('error', {'message': f"AI Interview Prep Error: {prep_result['error']}"}, room=sid)
            return

        print("Emitting interview_prep_materials to client.")
        await sio.emit('interview_prep_materials', prep_result, room=sid)

    except Exception as e:
        print(f"ERROR: Exception during request_interview_prep: {e}")
        traceback.print_exc()
        await sio.emit('error', {'message': f'Server error during interview prep: {e}'}, room=sid)

@sio.event
async def start_interview_simulation(sid, data):
    """
    Generates a personalized behavioral interview question based on resume and job description.
    """
    print(f"Received start_interview_simulation event from {sid}")
    resume_text = data.get('resume_text')
    job_description = data.get('job_description')

    if not resume_text or not job_description:
        await sio.emit('error', {'message': 'Resume text and Job Description are required for interview simulation.'}, room=sid)
        return

    try:
        simulation_prompt = f"""
As an expert interviewer. Based on the following resume and job description, generate ONE challenging behavioral interview question.
The question should focus on a skill or experience that is crucial for the role but might be a subtle gap or a strong point to elaborate on from the resume.
Frame it as a STAR method question (e.g., "Tell me about a time when...").
Also, provide a brief (1-2 sentences) context or scenario related to the question.

**IMPORTANT:** Your entire response MUST be a SINGLE JSON object. Do NOT include any other text, explanations, or markdown formatting.
The JSON object MUST have the following keys:
-   `question`: The behavioral interview question (string).
-   `scenario`: A brief contextual scenario for the question (string).
If you cannot generate valid JSON, provide the literal string '{JSON_ERROR_RESPONSE_LITERAL}'.

Resume:
```
{resume_text}
```

Job Description:
```
{job_description}
```
"""
        simulation_schema = {
            "type": "OBJECT",
            "properties": {
                "question": {"type": "STRING"},
                "scenario": {"type": "STRING"}
            },
            "required": ["question", "scenario"]
        }

        # Call LM Studio API for interview simulation question
        print(f"Calling {GLOBAL_MODEL_NAME} via LM Studio for interview simulation question...")
        simulation_result = await call_lm_studio_api(simulation_prompt, simulation_schema)

        if "error" in simulation_result:
            await sio.emit('error', {'message': f"AI Interview Simulation Error: {simulation_result['error']}"}, room=sid)
            return

        await sio.emit('interview_question', simulation_result, room=sid)

    except Exception as e:
        print(f"ERROR: Exception during start_interview_simulation: {e}")
        traceback.print_exc()
        await sio.emit('error', {'message': f'Server error during simulation: {e}'}, room=sid)

@sio.event
async def get_interview_feedback(sid, data):
    """
    Provides real-time feedback on a user's interview answer.
    """
    print(f"Received get_interview_feedback event from {sid}")
    question = data.get('question')
    user_answer = data.get('user_answer')
    job_description = data.get('job_description') # For context
    resume_text = data.get('resume_text') # For context

    if not question or not user_answer or not job_description or not resume_text:
        await sio.emit('error', {'message': 'Question, user answer, job description, and resume are required for feedback.'}, room=sid)
        return

    try:
        feedback_prompt = f"""
As an expert interview coach. Analyze the following user's interview answer for the given question, considering the job description and their resume.
Provide constructive feedback focusing on:
1.  **STAR Method Adherence:** Does the answer follow the Situation, Task, Action, Result structure? Point out missing components.
2.  **Relevance & Impact:** How well does the answer relate to the job description and demonstrate quantifiable impact? Suggest ways to improve relevance and add metrics.
3.  **Clarity & Conciseness:** Is the answer clear, easy to understand, and to the point?
4.  **Overall Suggestion:** A final overall tip for improving the answer.

**IMPORTANT:** Your entire response MUST be a SINGLE JSON object. Do NOT include any other text, explanations, or markdown formatting.
The JSON object MUST have the following keys:
-   `star_feedback`: string
-   `relevance_impact_feedback`: string
-   `clarity_feedback`: string
-   `overall_suggestion`: string
If you cannot generate valid JSON, provide the literal string '{JSON_ERROR_RESPONSE_LITERAL}'.

---
Job Description:
```
{job_description}
```

Resume:
```
{resume_text}
```

Interview Question:
```
{question}
```

User's Answer:
```
{user_answer}
```
---
"""
        feedback_schema = {
            "type": "OBJECT",
            "properties": {
                "star_feedback": {"type": "STRING"},
                "relevance_impact_feedback": {"type": "STRING"},
                "clarity_feedback": {"type": "STRING"},
                "overall_suggestion": {"type": "STRING"}
            },
            "required": ["star_feedback", "relevance_impact_feedback", "clarity_feedback", "overall_suggestion"]
        }

        # Call LM Studio API for interview feedback
        print(f"Calling {GLOBAL_MODEL_NAME} via LM Studio for interview feedback...")
        feedback_result = await call_lm_studio_api(feedback_prompt, feedback_schema)

        if "error" in feedback_result:
            await sio.emit('error', {'message': f"AI Feedback Error: {feedback_result['error']}"}, room=sid)
            return

        await sio.emit('interview_feedback', feedback_result, room=sid)

    except Exception as e:
        print(f"ERROR: Exception during get_interview_feedback: {e}")
        traceback.print_exc()
        await sio.emit('error', {'message': f'Server error during feedback: {e}'}, room=sid)

@sio.event
async def request_career_roadmap(sid, data):
    """
    Generates a predictive career trajectory and skill pathing roadmap.
    """
    print(f"Received request_career_roadmap event from {sid}")
    resume_text = data.get('resume_text')
    job_description = data.get('job_description') # Current target job
    desired_roles = data.get('desired_roles') # e.g., ["Senior Software Engineer", "Tech Lead"]

    if not resume_text or not job_description or not desired_roles:
        await sio.emit('error', {'message': 'Resume text, current job description, and desired roles are required for career roadmap.'}, room=sid)
        return

    # Join desired roles into a readable list for the prompt
    desired_roles_str = ", ".join(desired_roles)

    try:
        roadmap_prompt = f"""
As an expert career coach and talent development specialist.
Based on the provided resume and the current target job description, and the user's desired future career roles ({desired_roles_str}), generate a personalized career roadmap.

Your roadmap should:
1.  **Identify Key Skill Gaps:** What specific skills or experiences are missing from the resume for the user to progress towards "{desired_roles_str}"? Categorize these (e.g., Technical, Leadership, Communication, etc.).
2.  **Suggest Targeted Learning:** For each identified skill gap, recommend 1-2 specific types of online courses, certifications, or learning resources (e.g., "Advanced Python course on Coursera," "PMP Certification," "Public Speaking workshop").
3.  **Recommend Practical Projects/Experiences:** Suggest specific types of projects or work experiences (personal or professional) that would help bridge these gaps and make the resume more compelling for the desired roles.
4.  **Networking/Soft Skill Advice:** Provide 2-3 actionable tips related to networking, mentorship, or developing crucial soft skills for career advancement.
5.  **Aspirational Summary/Next Steps:** A brief summary of how their resume could look after implementing these steps, and a concluding encouraging remark.

**IMPORTANT:** Your entire response MUST be a SINGLE JSON object. Do NOT include any other text, explanations, or markdown formatting.
The JSON object MUST have the following keys:
-   `skill_gaps`: An array of objects, each with `category` (string) and `skills` (array of strings).
-   `learning_recommendations`: An array of objects, each with `skill_area` (string) and `resources` (array of strings, e.g., "Course: AWS Certified Solutions Architect - Associate (Coursera)").
-   `project_ideas`: An array of strings, each being a project idea.
-   `networking_tips`: An array of strings, each being a networking/soft skill tip.
-   `aspirational_summary`: string
If you cannot generate valid JSON, provide the literal string '{JSON_ERROR_RESPONSE_LITERAL}'.

---
Resume:
```
{resume_text}
```

Current Job Description:
```
{job_description}
```

Desired Future Roles:
```
{desired_roles_str}
```
---
"""
        roadmap_schema = {
            "type": "OBJECT",
            "properties": {
                "skill_gaps": {
                    "type": "ARRAY",
                    "items": {
                        "type": "OBJECT",
                        "properties": {
                            "category": {"type": "STRING"},
                            "skills": {"type": "ARRAY", "items": {"type": "STRING"}}
                        },
                        "required": ["category", "skills"]
                    }
                },
                "learning_recommendations": {
                    "type": "ARRAY",
                    "items": {
                        "type": "OBJECT",
                        "properties": {
                            "skill_area": {"type": "STRING"},
                            "resources": {"type": "ARRAY", "items": {"type": "STRING"}}
                        },
                        "required": ["skill_area", "resources"]
                    }
                },
                "project_ideas": {"type": "ARRAY", "items": {"type": "STRING"}},
                "networking_tips": {"type": "ARRAY", "items": {"type": "STRING"}},
                "aspirational_summary": {"type": "STRING"}
            },
            "required": ["skill_gaps", "learning_recommendations", "project_ideas", "networking_tips", "aspirational_summary"]
        }

        # Call LM Studio API for career roadmap
        print(f"Calling {GLOBAL_MODEL_NAME} via LM Studio for career roadmap...")
        roadmap_result = await call_lm_studio_api(roadmap_prompt, roadmap_schema)

        if "error" in roadmap_result:
            await sio.emit('error', {'message': f"AI Career Roadmap Error: {roadmap_result['error']}"}, room=sid)
            return

        await sio.emit('career_roadmap', roadmap_result, room=sid)

    except Exception as e:
        print(f"ERROR: Exception during request_career_roadmap: {e}")
        traceback.print_exc()
        await sio.emit('error', {'message': f'Server error during career roadmap: {e}'}, room=sid)


# --- Main entry point ---
if __name__ == '__main__':
    # To run locally, you would typically install required packages:
    # pip install fastapi uvicorn python-socketio PyMuPDF reportlab httpx python-dotenv python-docx openpyxl
    # And then ensure LM Studio is running and serving 'gemma3:4b' on http://192.168.234.1:1234
    print("Starting FastAPI-Socket.IO server with Uvicorn...")
    uvicorn.run(app_with_sio, host='0.0.0.0', port=5000, log_level="info")

